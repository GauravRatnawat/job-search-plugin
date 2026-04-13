#!/usr/bin/env python3
"""Excel job tracker — save, view, update jobs. Returns JSON.

Usage:
    python scripts/tracker.py save '[{...job JSON...}]'
    python scripts/tracker.py view [--status Applied]
    python scripts/tracker.py update <job_id> <status> [--notes "..."]
    python scripts/tracker.py summary
"""

import argparse
import json
import os
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from filelock import FileLock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TRACKER_PATH = Path(os.environ.get("TRACKER_PATH", "./job_tracker.xlsx"))
LOCK_PATH = TRACKER_PATH.with_suffix(".lock")

COLUMNS = [
    "ID",
    "Company",
    "Title",
    "Location",
    "Remote",
    "Score",
    "Grade",
    "Status",
    "URL",
    "Source",
    "Date Added",
    "Date Applied",
    "Notes",
    "Salary",
]

VALID_STATUSES = {
    "New",
    "Reviewing",
    "Applied",
    "Interviewing",
    "Rejected",
    "Offer",
    "Archived",
}

STATUS_COLORS = {
    "New": "CCE5FF",
    "Reviewing": "FFF3CD",
    "Applied": "D4EDDA",
    "Interviewing": "D1ECF1",
    "Rejected": "F8D7DA",
    "Offer": "C3E6CB",
    "Archived": "E2E3E5",
}

GRADE_COLORS = {
    "A": "28A745",
    "B": "6F9E3A",
    "C": "FFC107",
    "D": "FD7E14",
    "F": "DC3545",
}

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Prevent CSV/formula injection in Excel cells."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return f"'{value}"
    return value


def _score_to_grade(score):
    """Convert a numeric score (0-100) to a letter grade."""
    if not isinstance(score, (int, float)):
        return ""
    if score >= 85:
        return "A"
    if score >= 70:
        return "B"
    if score >= 55:
        return "C"
    if score >= 40:
        return "D"
    return "F"


def ensure_workbook():
    if TRACKER_PATH.exists():
        return load_workbook(str(TRACKER_PATH))
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Tracker"
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    for i, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = {
        "A": 14,
        "B": 20,
        "C": 30,
        "D": 20,
        "E": 8,
        "F": 8,
        "G": 8,
        "H": 14,
        "I": 40,
        "J": 12,
        "K": 14,
        "L": 14,
        "M": 30,
        "N": 20,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return wb


def find_row(ws, job_id):
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == job_id:
            return row[0].row
    return None


def format_row(ws, row_num):
    status = str(ws.cell(row=row_num, column=8).value or "")
    grade = str(ws.cell(row=row_num, column=7).value or "")
    if status in STATUS_COLORS:
        ws.cell(row=row_num, column=8).fill = PatternFill(
            start_color=STATUS_COLORS[status],
            end_color=STATUS_COLORS[status],
            fill_type="solid",
        )
    if grade in GRADE_COLORS:
        ws.cell(row=row_num, column=7).font = Font(bold=True, color=GRADE_COLORS[grade])


def row_to_dict(row):
    job = {}
    for i, col in enumerate(COLUMNS):
        job[col] = row[i] if i < len(row) else ""
    return job


def safe_save(wb):
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(TRACKER_PATH.parent))
    os.close(fd)
    try:
        wb.save(tmp)
        shutil.move(tmp, str(TRACKER_PATH))
    except Exception:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise


def salary_str(sal_min, sal_max, currency="USD"):
    if sal_min and sal_max:
        return f"{currency} {sal_min:,.0f} - {sal_max:,.0f}"
    if sal_min:
        return f"{currency} {sal_min:,.0f}+"
    if sal_max:
        return f"Up to {currency} {sal_max:,.0f}"
    return ""


# ── Commands ────────────────────────────────────────────────────


def cmd_save(jobs_json):
    """Save jobs to tracker. Input: JSON array of job dicts."""
    try:
        jobs = json.loads(jobs_json) if isinstance(jobs_json, str) else jobs_json
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"Invalid JSON: {e}"}))
        return

    lock = FileLock(str(LOCK_PATH), timeout=10)
    added = updated = 0

    with lock:
        wb = ensure_workbook()
        ws = wb.active
        for job in jobs:
            jid = job.get("id", "")
            if not jid:
                # Auto-generate an ID from company + title
                company = job.get("company", "unknown")
                title = job.get("title", job.get("role", "unknown"))
                slug = f"{company}-{title}".lower()
                slug = slug.replace(" ", "-").replace("/", "-")[:60]
                jid = slug
            existing = find_row(ws, jid)
            if existing:
                score = job.get("score", job.get("match_score", ""))
                grade = job.get("grade", "") or _score_to_grade(score)
                ws.cell(row=existing, column=6, value=score)
                ws.cell(row=existing, column=7, value=grade)
                format_row(ws, existing)
                updated += 1
            else:
                sal = salary_str(job.get("salary_min"), job.get("salary_max"))
                score = job.get("score", job.get("match_score", ""))
                grade = job.get("grade", "") or _score_to_grade(score)
                ws.append(
                    [
                        _sanitize_cell(jid),
                        _sanitize_cell(job.get("company", "")),
                        _sanitize_cell(job.get("title", job.get("role", ""))),
                        _sanitize_cell(job.get("location", "")),
                        "Yes" if job.get("remote") else "No",
                        job.get("score", job.get("match_score", "")),
                        grade,
                        job.get("status", "New"),
                        _sanitize_cell(job.get("url", "")),
                        _sanitize_cell(job.get("source", "")),
                        datetime.now().strftime("%Y-%m-%d"),
                        "",
                        _sanitize_cell(job.get("notes", "")),
                        _sanitize_cell(sal),
                    ]
                )
                format_row(ws, ws.max_row)
                added += 1
        safe_save(wb)

    print(
        json.dumps(
            {
                "message": f"{added} added, {updated} updated",
                "added": added,
                "updated": updated,
                "file": str(TRACKER_PATH),
            },
            indent=2,
        )
    )


def cmd_view(status_filter=""):
    """View tracked jobs."""
    if not TRACKER_PATH.exists():
        print(json.dumps({"total": 0, "jobs": [], "message": "No tracker file yet."}))
        return

    lock = FileLock(str(LOCK_PATH), timeout=10)
    with lock:
        wb = load_workbook(str(TRACKER_PATH), read_only=True)
        ws = wb.active
        jobs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = row_to_dict(row)
            if status_filter and d.get("Status") != status_filter:
                continue
            jobs.append(d)
        wb.close()

    print(
        json.dumps(
            {
                "total": len(jobs),
                "filter": status_filter or "all",
                "jobs": jobs,
            },
            indent=2,
            default=str,
        )
    )


def get_job_record(job_id):
    """Return a tracker row as a dict, or None if not found."""
    if not TRACKER_PATH.exists():
        return None

    lock = FileLock(str(LOCK_PATH), timeout=10)
    with lock:
        wb = load_workbook(str(TRACKER_PATH), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == job_id:
                job = row_to_dict(row)
                wb.close()
                return job
        wb.close()
    return None


def update_job_status(job_id, status, notes=""):
    """Update a job's status and return a JSON-serializable result."""
    if status not in VALID_STATUSES:
        raise ValueError(f"Invalid status '{status}'. Valid: {sorted(VALID_STATUSES)}")

    lock = FileLock(str(LOCK_PATH), timeout=10)
    with lock:
        wb = ensure_workbook()
        ws = wb.active
        row = find_row(ws, job_id)
        if not row:
            raise KeyError(f"Job ID '{job_id}' not found")

        ws.cell(row=row, column=8, value=status)
        if status == "Applied":
            ws.cell(row=row, column=12, value=datetime.now().strftime("%Y-%m-%d"))
        if notes:
            existing = ws.cell(row=row, column=13).value or ""
            safe_notes = _sanitize_cell(f"{existing}; {notes}" if existing else notes)
            ws.cell(row=row, column=13, value=safe_notes)
        format_row(ws, row)
        safe_save(wb)

    return {"message": f"Job {job_id} → {status}", "notes": notes or "(none)"}


def cmd_update(job_id, status, notes=""):
    """Update a job's status."""
    try:
        print(json.dumps(update_job_status(job_id, status, notes)))
    except ValueError as exc:
        print(json.dumps({"error": str(exc)}))
    except KeyError as exc:
        print(json.dumps({"error": str(exc)}))


def cmd_summary():
    """Pipeline summary."""
    if not TRACKER_PATH.exists():
        print(json.dumps({"total": 0, "by_status": {}, "by_grade": {}, "avg_score": 0}))
        return

    lock = FileLock(str(LOCK_PATH), timeout=10)
    with lock:
        wb = load_workbook(str(TRACKER_PATH), read_only=True)
        ws = wb.active
        by_status = {}
        by_grade = {}
        scores = []
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            total += 1
            status = row[7] if len(row) > 7 else "Unknown"
            grade = row[6] if len(row) > 6 else "?"
            score = row[5] if len(row) > 5 else None
            by_status[status] = by_status.get(status, 0) + 1
            by_grade[grade] = by_grade.get(grade, 0) + 1
            if score and isinstance(score, (int, float)):
                scores.append(float(score))
        wb.close()

    print(
        json.dumps(
            {
                "total": total,
                "by_status": by_status,
                "by_grade": by_grade,
                "avg_score": round(sum(scores) / len(scores), 1) if scores else 0,
            },
            indent=2,
        )
    )


# ── CLI ─────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Job tracker (Excel)")
    sub = parser.add_subparsers(dest="command", required=True)

    p_save = sub.add_parser("save", help="Save jobs to tracker")
    p_save.add_argument("jobs_json", help="JSON array of job objects")

    p_view = sub.add_parser("view", help="View tracked jobs")
    p_view.add_argument("--status", default="", help="Filter by status")

    p_update = sub.add_parser("update", help="Update job status")
    p_update.add_argument("job_id", help="Job ID")
    p_update.add_argument("status", help="New status")
    p_update.add_argument("--notes", default="", help="Notes to append")

    sub.add_parser("summary", help="Pipeline summary stats")

    args = parser.parse_args()

    if args.command == "save":
        cmd_save(args.jobs_json)
    elif args.command == "view":
        cmd_view(args.status)
    elif args.command == "update":
        cmd_update(args.job_id, args.status, getattr(args, "notes", ""))
    elif args.command == "summary":
        cmd_summary()


if __name__ == "__main__":
    main()
